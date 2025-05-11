from openpyxl import load_workbook

workbook = load_workbook("sample2.xlsx")

sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
  # print(row)
  # for cell in row:
  #   print(cell, end=" ")
  row = " !! ".join(row)
  # print()