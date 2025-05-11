import openpyxl
import os

result_workbook = openpyxl.Workbook()

result_sheet = result_workbook.active

file_list = os.listdir()

for file in file_list:
  if file.endswith("xlsx"):
    print(file)
    workbook = openpyxl.load_workbook(file)

    sheet = workbook.worksheets[0]

    for row in sheet.iter_rows(values_only=True):
      # print(row)
      result_sheet.append(row)

result_workbook.save("result.xlsx")