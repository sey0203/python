import openpyxl

workbook = openpyxl.Workbook()

sheet = workbook.active

data_dict = {"이름": "김플", "성별": "남자", "직업": "프로그래머"}

row_num = 1
for data in data_dict.items():
  print(data[0])
  print(data[1])
  sheet.cell(row=row_num, column=1, value=data[0])
  sheet.cell(row=row_num, column=2, value=data[1])

  row_num += 1

workbook.save("sample2.xlsx")