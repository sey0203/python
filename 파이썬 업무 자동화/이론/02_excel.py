import openpyxl
import os

# workbook = openpyxl.Workbook()

# sheet = workbook.active

# sheet["A1"] = "Hello"

# for i in range(1, 10):
#   sheet.cell(row=i, column=1, value="Hi!")

# workbook.save("hello.xlsx")

file_list = os.listdir()

for file in file_list:
  if file.endswith("xlsx"):

    workbook = openpyxl.load_workbook("hello.xlsx")

    sheet = workbook.worksheets[0]

    cell = sheet["A1"]

    print(cell.value)